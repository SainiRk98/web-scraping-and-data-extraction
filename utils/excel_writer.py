"""
utils/excel_writer.py
---------------------
Centralised Excel writing helpers built on top of pandas + openpyxl.
Every task calls one of these functions to persist its data.
"""

from pathlib import Path
from typing import List, Dict, Any, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from utils.logger import get_logger

log = get_logger(__name__)

# Header style constants
HEADER_FONT  = Font(bold=True, color="FFFFFF")
HEADER_FILL  = PatternFill("solid", fgColor="2E75B6")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_worksheet(ws) -> None:
    """Apply header styling and auto-fit column widths."""
    for cell in ws[1]:
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)


def write_dataframe(df: pd.DataFrame, path: Path, sheet_name: str = "Sheet1") -> None:
    """
    Write a single DataFrame to an Excel file.
    If the file already exists the sheet is replaced; other sheets are kept.
    """
    path = Path(path)
    try:
        if path.exists():
            with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Apply styling
        wb = load_workbook(path)
        _style_worksheet(wb[sheet_name])
        wb.save(path)
        log.info("Saved %d rows → %s [%s]", len(df), path.name, sheet_name)
    except Exception as exc:
        log.error("Failed to write Excel %s: %s", path, exc)
        raise


def write_multiple_sheets(sheets: Dict[str, pd.DataFrame], path: Path) -> None:
    """
    Write multiple DataFrames, each to its own sheet, into one Excel file.
    Used by the PDF extractor (one sheet per page).
    """
    path = Path(path)
    try:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for sheet_name, df in sheets.items():
                safe_name = sheet_name[:31]  # Excel sheet name limit
                df.to_excel(writer, sheet_name=safe_name, index=False)

        # Style every sheet
        wb = load_workbook(path)
        for sheet_name in wb.sheetnames:
            _style_worksheet(wb[sheet_name])
        wb.save(path)
        log.info("Saved %d sheets → %s", len(sheets), path.name)
    except Exception as exc:
        log.error("Failed to write multi-sheet Excel %s: %s", path, exc)
        raise


def records_to_excel(records: List[Dict[str, Any]], path: Path, sheet_name: str = "Sheet1") -> None:
    """Convenience wrapper: convert a list of dicts → DataFrame → Excel."""
    df = pd.DataFrame(records)
    if df.empty:
        log.warning("No records to write for %s.", path.name)
        df = pd.DataFrame(columns=["Note"])
        df.loc[0] = ["No data extracted"]
    write_dataframe(df, path, sheet_name=sheet_name)
